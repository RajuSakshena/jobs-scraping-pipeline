import os
import traceback
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment

from scrapers.estm.estm import scrape_jobs as scrape_estm_jobs
from scrapers.c40.c40 import scrape_c40_jobs
from scrapers.developmentaid.developmentaid import scrape_jobs as scrape_developmentaid_jobs
from scrapers.onepurpos import scrape_onepurpose_jobs  # ✅ NEW


# ======================================================
# PATH SETUP
# ======================================================

BASE_DIR = os.getcwd()
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
COMBINED_FILE = os.path.join(OUTPUT_DIR, "Combined.xlsx")

FINAL_COLUMNS = [
    "Source",
    "Title",
    "Description",
    "Matched_Vertical",
    "Deadline",
    "Apply_Link"
]


# ======================================================
# EXCEL FORMAT FUNCTION
# ======================================================

def format_excel(path):
    try:
        wb = load_workbook(path)
        ws = wb.active

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 55
        ws.column_dimensions["C"].width = 120
        ws.column_dimensions["D"].width = 35
        ws.column_dimensions["E"].width = 25
        ws.column_dimensions["F"].width = 60

        for row in ws.iter_rows(min_row=2):
            ws.row_dimensions[row[0].row].height = 95
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        wb.save(path)
        print("✅ Excel formatting applied")

    except Exception:
        print("❌ Excel formatting failed")
        traceback.print_exc()


# ======================================================
# CLEAN LINK FUNCTION
# ======================================================

def clean_link(link):
    if not isinstance(link, str) or not link.strip():
        return ""

    link = link.strip()

    if 'HYPERLINK(' in link:
        match = re.search(r'HYPERLINK\("([^"]+)"', link)
        if match:
            return match.group(1)

    return link


# ======================================================
# CLEAN DESCRIPTION FUNCTION
# ======================================================

def clean_description(desc):
    if not desc or str(desc).strip().lower() in ["", "nan", "none"]:
        return "No description available"
    return str(desc).strip()


# ======================================================
# MAIN RUNNER FUNCTION
# ======================================================

def run_all_scrapers_and_combine():
    try:
        print("🚀 Starting scraper process...")
        os.makedirs(OUTPUT_DIR, exist_ok=True)

        combined_rows = []

        # ======================================================
        # ESTM SCRAPER
        # ======================================================

        try:
            print("🔎 Running ESTM scraper...")
            estm_df = scrape_estm_jobs()

            if estm_df is not None and not estm_df.empty:
                for _, row in estm_df.iterrows():
                    combined_rows.append({
                        "Source": "ESTM",
                        "Title": row.get("Title"),
                        "Description": clean_description(row.get("Description")),
                        "Matched_Vertical": row.get("Matched_Vertical"),
                        "Deadline": row.get("Deadline"),
                        "Apply_Link": clean_link(row.get("Apply_Link"))
                    })

                print(f"✅ ESTM rows added: {len(estm_df)}")
            else:
                print("⚠ ESTM returned no data")

        except Exception:
            print("❌ ESTM failed")
            traceback.print_exc()


        # ======================================================
        # C40 SCRAPER
        # ======================================================

        try:
            print("🔎 Running C40 scraper...")
            c40_df = scrape_c40_jobs()

            if c40_df is not None and not c40_df.empty:
                for _, row in c40_df.iterrows():
                    combined_rows.append({
                        "Source": "C40",
                        "Title": row.get("Title"),
                        "Description": clean_description(row.get("Description")),
                        "Matched_Vertical": row.get("Matched_Vertical"),
                        "Deadline": row.get("Deadline"),
                        "Apply_Link": clean_link(row.get("Apply_Link"))
                    })

                print(f"✅ C40 rows added: {len(c40_df)}")
            else:
                print("⚠ C40 returned no data")

        except Exception:
            print("❌ C40 failed")
            traceback.print_exc()


        # ======================================================
        # ONEPURPOSE SCRAPER ✅ NEW
        # ======================================================

        try:
            print("🔎 Running OnePurpose scraper...")
            op_df = scrape_onepurpose_jobs()

            if op_df is not None and not op_df.empty:
                for _, row in op_df.iterrows():
                    combined_rows.append({
                        "Source": "onepurpos",
                        "Title": row.get("Title"),
                        "Description": clean_description(row.get("Description")),
                        "Matched_Vertical": row.get("Matched_Vertical"),
                        "Deadline": row.get("Deadline"),
                        "Apply_Link": clean_link(row.get("Apply_Link"))
                    })

                print(f"✅ OnePurpose rows added: {len(op_df)}")
            else:
                print("⚠ OnePurpose returned no data")

        except Exception:
            print("❌ OnePurpose failed")
            traceback.print_exc()


        # ======================================================
        # DevelopmentAid SCRAPER
        # ======================================================

        try:
            print("🔎 Running DevelopmentAid scraper...")
            da_df = scrape_developmentaid_jobs()

            if da_df is not None and not da_df.empty:
                for _, row in da_df.iterrows():
                    combined_rows.append({
                        "Source": "DevelopmentAid",
                        "Title": row.get("Title"),
                        "Description": clean_description(row.get("Description")),
                        "Matched_Vertical": row.get("Category"),
                        "Deadline": row.get("Deadline"),
                        "Apply_Link": clean_link(row.get("Apply_Link"))
                    })

                print(f"✅ DevelopmentAid rows added: {len(da_df)}")
            else:
                print("⚠ DevelopmentAid returned no data")

        except Exception:
            print("❌ DevelopmentAid failed")
            traceback.print_exc()


        # ======================================================
        # FINAL EXPORT
        # ======================================================

        if not combined_rows:
            print("❌ No data collected from any scraper")
            return None

        combined_df = pd.DataFrame(combined_rows, columns=FINAL_COLUMNS)

        combined_df.to_excel(COMBINED_FILE, index=False, engine="openpyxl")

        format_excel(COMBINED_FILE)

        print("📁 Output directory:", OUTPUT_DIR)
        print("📄 Combined file created at:", COMBINED_FILE)
        print("🎉 Scraping process completed successfully!")

        return COMBINED_FILE

    except Exception:
        print("🔥 Critical error in main runner")
        traceback.print_exc()
        return None


# ======================================================
# ENTRY POINT
# ======================================================

if __name__ == "__main__":
    run_all_scrapers_and_combine()
