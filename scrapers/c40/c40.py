import os
import time
import json
import re
import pandas as pd
from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# ======================================================
# PATH SETUP
# ======================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.abspath(os.path.join(BASE_DIR, "..", ".."))

CAREERS_URL = "https://c40.bamboohr.com/careers"
BASE_URL = "https://c40.bamboohr.com"

KEYWORDS_FILE = os.path.join(BASE_DIR, "keywords.json")
OUTPUT_DIR = os.path.join(PROJECT_ROOT, "output")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "c40_jobs.xlsx")


def load_keywords():
    with open(KEYWORDS_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def match_verticals(title, description, keywords):
    text = f"{title} {description}".lower()
    matched = []
    for vertical, words in keywords.items():
        for w in words:
            if re.search(rf"\b{re.escape(w.lower())}\b", text):
                matched.append(vertical)
                break
    return ", ".join(matched) if matched else "N/A"


def scrape_c40_jobs():
    keywords = load_keywords()
    jobs = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()

        page.goto(CAREERS_URL, timeout=60000)
        page.wait_for_load_state("networkidle")
        time.sleep(2)

        soup = BeautifulSoup(page.content(), "html.parser")

        job_links = []
        for a in soup.select("a[href^='/careers/']"):
            href = a.get("href")
            title = a.get_text(strip=True)
            if href and title:
                job_links.append((BASE_URL + href, title))

        job_links = list(dict.fromkeys(job_links))

        for job_url, fallback_title in job_links:
            page.goto(job_url, timeout=60000)
            page.wait_for_load_state("networkidle")
            time.sleep(2)

            job_soup = BeautifulSoup(page.content(), "html.parser")

            title_tag = job_soup.find("h3", class_="fabric-oxx0vk-root")
            title = title_tag.get_text(strip=True) if title_tag else fallback_title

            desc_blocks = job_soup.find_all("div", class_="fabric-95l02p-description")
            description = "\n".join(d.get_text(strip=True) for d in desc_blocks)

            if not description:
                continue

            matched_vertical = match_verticals(title, description, keywords)

            jobs.append({
                "Title": title,
                "Description": description,
                "Matched_Vertical": matched_vertical,
                "Apply_Link": f'=HYPERLINK("{job_url}", "{title.replace(chr(34), "")}")'
            })

        browser.close()

    if not jobs:
        print("❌ No C40 jobs extracted")
        return pd.DataFrame()

    df = pd.DataFrame(
        jobs,
        columns=["Title", "Description", "Matched_Vertical", "Apply_Link"]
    )

    # Save Excel (optional output)
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    df.to_excel(OUTPUT_FILE, index=False, engine="openpyxl")

    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active

    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 120
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 50

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(OUTPUT_FILE)

    print(f"✅ C40 Excel created: {OUTPUT_FILE}")

    return df   # 🔥 THIS IS THE KEY FIX


if __name__ == "__main__":
    scrape_c40_jobs()
